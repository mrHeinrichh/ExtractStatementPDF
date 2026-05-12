from __future__ import annotations

import copy
import csv
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from pypdf import PdfReader


MONTH_DIR_RE = re.compile(r"^\d{2}-")
MONTH_TOKEN_RE = re.compile(r"^(0[1-9]|1[0-2])\d{4}$")
AR_LINE_RE = re.compile(
    r"(\d{2}-(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)-\d{2})\s+(\d+)\s+([\d,.]+)\s+([\d,.]+)\s+([\d,.]+)"
)
NUMERIC_REF_RE = re.compile(r"^\d+$")
EXCLUDED_DIRS = {"test", "incomplete", "investgiate", "investigate", "new folder"}
EXCLUDED_FILES = {"SOA Results.xlsx", "RxOffice Accounting Issues.xlsx"}
INSERT_HEADERS = [
    "Rx Gross",
    "Rx Discount",
    "Rx NET",
    "AR Gross",
    "AR Discount",
    "AR NET",
    "Variance",
]


@dataclass(frozen=True)
class OrderSnapshot:
    rx_gross: float | None
    rx_discount: float | None
    rx_net: float | None
    ar_gross: float | None
    ar_discount: float | None
    ar_net: float | None
    variance: float


def build_source_lookup(soa_root: Path) -> dict[str, OrderSnapshot]:
    lookup: dict[str, OrderSnapshot] = {}
    for month_dir in sorted(
        (path for path in soa_root.iterdir() if path.is_dir() and MONTH_DIR_RE.match(path.name)),
        key=lambda p: p.name.lower(),
    ):
        files = get_month_source_files(month_dir)
        csv_files = [path for path in files if path.suffix.lower() == ".csv"]
        ar_files = [path for path in files if path.suffix.lower() in {".pdf", ".xls", ".xlsx"}]
        for csv_file in csv_files:
            base_key = build_source_lookup_key(csv_file)
            matching_ar = [path for path in ar_files if build_source_lookup_key(path) == base_key]
            if not matching_ar:
                continue

            rx_orders = parse_rx_csv(csv_file)
            ar_orders = parse_ar_orders(matching_ar)
            month_label = extract_month_label(csv_file)
            customer_key = base_key

            for reference in sorted(set(rx_orders) | set(ar_orders)):
                if not NUMERIC_REF_RE.fullmatch(reference):
                    continue
                rx = rx_orders.get(reference, {})
                ar = ar_orders.get(reference, {})
                rx_net = rx.get("net")
                ar_net = ar.get("net")
                variance = abs((ar_net or 0.0) - (rx_net or 0.0))
                lookup[f"{month_label}|{customer_key}|{reference}"] = OrderSnapshot(
                    rx.get("gross"),
                    rx.get("discount"),
                    rx_net,
                    ar.get("gross"),
                    ar.get("discount"),
                    ar_net,
                    variance,
                )
    return lookup


def get_month_source_files(month_dir: Path) -> list[Path]:
    files: list[Path] = []
    for path in month_dir.rglob("*"):
        if not path.is_file():
            continue
        if path.name in EXCLUDED_FILES:
            continue
        if path.suffix.lower() not in {".csv", ".pdf", ".xls", ".xlsx"}:
            continue
        if any(part.lower() in EXCLUDED_DIRS for part in path.relative_to(month_dir).parts[:-1]):
            continue
        files.append(path)
    return files


def parse_rx_csv(path: Path) -> dict[str, dict[str, float]]:
    orders: dict[str, dict[str, float]] = {}
    with path.open("r", newline="", encoding="utf-8-sig") as handle:
        reader = csv.reader(handle)
        next(reader, None)
        for row in reader:
            if len(row) < 16:
                continue
            reference = row[12].strip()
            if not NUMERIC_REF_RE.fullmatch(reference):
                continue
            orders[reference] = {
                "gross": parse_decimal(row[13]),
                "discount": parse_decimal(row[14]),
                "net": parse_decimal(row[15]),
            }
    return orders


def parse_ar_orders(paths: Iterable[Path]) -> dict[str, dict[str, float]]:
    orders: dict[str, dict[str, float]] = {}
    for path in paths:
        if path.suffix.lower() == ".pdf":
            texts = []
            reader = PdfReader(str(path))
            for page in reader.pages:
                text = page.extract_text() or ""
                texts.append(text)
            content = "\n".join(texts)
            for match in AR_LINE_RE.finditer(content):
                orders[match.group(2)] = {
                    "gross": parse_decimal(match.group(3)),
                    "discount": parse_decimal(match.group(4)),
                    "net": parse_decimal(match.group(5)),
                }
    return orders


def build_source_lookup_key(path: Path) -> str:
    name = path.stem.strip()
    name = re.sub(r"_P\d+$", "", name, flags=re.IGNORECASE)
    name = re.sub(r"\s+\d{1,2}-\d{1,2}-\d{1,2},\s*\d{4}$", "", name, flags=re.IGNORECASE)
    name = re.sub(r"_(?:\d+(?:st|nd|rd|th))$", "", name, flags=re.IGNORECASE)
    parts = [part.strip() for part in name.split("_") if part.strip()]
    month_index = next((index for index, value in enumerate(parts) if MONTH_TOKEN_RE.fullmatch(value)), -1)
    if month_index > 0:
        name = "_".join(parts[:month_index])
    return normalize_customer_key(name)


def extract_month_label(path: Path) -> str:
    parts = [part.strip() for part in path.stem.split("_") if part.strip()]
    for part in parts:
        if MONTH_TOKEN_RE.fullmatch(part):
            return datetime.strptime(part, "%m%Y").strftime("%B %Y")
    for part in path.parts:
        if MONTH_DIR_RE.match(part):
            month_name = part.split("-", 1)[1]
            for fmt in ("%B %Y", "%B"):
                try:
                    parsed = datetime.strptime(month_name, fmt)
                    year = parsed.year if fmt == "%B %Y" else 2025
                    return datetime(year, parsed.month, 1).strftime("%B %Y")
                except ValueError:
                    pass
    raise ValueError(f"Could not determine month from {path}")


def normalize_customer_key(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9]+", "", value or "").lower()


def parse_decimal(value: str | None) -> float:
    text = (value or "").strip()
    text = text.replace("PHP", "").replace(",", "").replace('"', "").strip()
    if not text:
        return 0.0
    return float(text)


def normalize_month_label(value: object) -> str:
    if isinstance(value, datetime):
        return value.strftime("%B %Y")
    text = ("" if value is None else str(value)).strip()
    if not text:
        return ""
    for fmt in ("%m%Y", "%B %Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).strftime("%B %Y")
        except ValueError:
            pass
    try:
        return datetime.fromisoformat(text).strftime("%B %Y")
    except ValueError:
        return text


def write_headers(ws) -> None:
    for offset, header in enumerate(INSERT_HEADERS, start=4):
        ws.cell(1, offset).value = header


def add_columns_and_fill(workbook_path: Path, lookup: dict[str, OrderSnapshot], output_root: Path) -> tuple[list[Path], dict[str, list[int]], int]:
    wb = load_workbook(workbook_path)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb[wb.sheetnames[0]]
    has_headers = ws.cell(1, 4).value == "Rx Gross" and ws.cell(1, 10).value == "Variance"
    if not has_headers:
        next_data_col = next((col for col in range(4, ws.max_column + 1) if ws.cell(1, col).value is not None), 4)
        blank_slots = max(0, next_data_col - 4)
        missing_slots = max(0, 7 - blank_slots)
        if missing_slots:
            ws.insert_cols(next_data_col, missing_slots)
        write_headers(ws)
        style_source_col = next_data_col + missing_slots
        for col in range(4, 11):
            ws.column_dimensions[ws.cell(1, col).column_letter].width = ws.column_dimensions[
                ws.cell(1, style_source_col).column_letter
            ].width
            for row in range(1, ws.max_row + 1):
                source = ws.cell(row, style_source_col)
                target = ws.cell(row, col)
                target._style = copy.copy(source._style)
                if source.number_format:
                    target.number_format = source.number_format
        for col in range(4, 11):
            for row in range(2, ws.max_row + 1):
                ws.cell(row, col).number_format = "#,##0.00"
    else:
        for row in range(2, ws.max_row + 1):
            for col in range(4, 11):
                ws.cell(row, col).value = None

    groups: dict[str, list[int]] = defaultdict(list)
    unmatched = 0
    for row in range(2, ws.max_row + 1):
        reference = str(ws.cell(row, 3).value or "").strip()
        if reference.endswith(".0"):
            reference = reference[:-2]
        if not NUMERIC_REF_RE.fullmatch(reference):
            continue
        month_label = normalize_month_label(ws.cell(row, 2).value)
        if not month_label:
            continue
        customer_key = normalize_customer_key(str(ws.cell(row, 1).value or ""))
        key = f"{month_label}|{customer_key}|{reference}"
        snapshot = lookup.get(key)
        if snapshot is None:
            unmatched += 1
        else:
            ws.cell(row, 4).value = snapshot.rx_gross
            ws.cell(row, 5).value = snapshot.rx_discount
            ws.cell(row, 6).value = snapshot.rx_net
            ws.cell(row, 7).value = snapshot.ar_gross
            ws.cell(row, 8).value = snapshot.ar_discount
            ws.cell(row, 9).value = snapshot.ar_net
            ws.cell(row, 10).value = snapshot.variance
        groups[month_label].append(row)

    output_root.mkdir(parents=True, exist_ok=True)
    master_path = output_root / "RxOffice Accounting Issues - Enriched.xlsx"
    wb.save(master_path)

    output_files = [master_path]
    for month_label in sorted(groups, key=lambda label: datetime.strptime(label, "%B %Y")):
        month_wb = load_workbook(master_path)
        month_ws = month_wb["Sheet1"] if "Sheet1" in month_wb.sheetnames else month_wb[month_wb.sheetnames[0]]
        keep = set(groups[month_label])
        for row in range(month_ws.max_row, 1, -1):
            if row not in keep:
                month_ws.delete_rows(row)
        month_path = output_root / f"RxOffice Accounting Issues - {month_label}.xlsx"
        month_wb.save(month_path)
        output_files.append(month_path)
    return output_files, groups, unmatched


def main() -> None:
    import sys

    soa_root = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(r"G:\.shortcut-targets-by-id\1VLaMf4DXM2APBuWDeKAJW_sNUkXgk7tk\Plastilens\05 - Testing\SOA")
    workbook_path = Path(sys.argv[2]) if len(sys.argv) > 2 else soa_root / "RxOffice Accounting Issues.xlsx"
    output_root = Path(sys.argv[3]) if len(sys.argv) > 3 else Path.cwd() / "outputs" / "accounting-issues-split"

    lookup = build_source_lookup(soa_root)
    output_files, groups, unmatched = add_columns_and_fill(workbook_path, lookup, output_root)

    print(f"lookup_entries={len(lookup)}")
    print(f"months={len(groups)}")
    print(f"unmatched_numeric_rows={unmatched}")
    for month_label in sorted(groups, key=lambda label: datetime.strptime(label, '%B %Y')):
        print(f"{month_label}: {len(groups[month_label])}")
    for path in output_files:
        print(path)


if __name__ == "__main__":
    main()
